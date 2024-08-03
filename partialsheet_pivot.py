def main():
    # 기존 코드

    import pandas as pd
    import openpyxl

    # 중간 파일 경로
    intermediate_file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-rmpw.xlsx'
    # 최종 저장 파일 경로
    final_file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-partial_pivot.xlsx'

    # Load the third sheet data
    df = pd.read_excel(intermediate_file_path, sheet_name=2, usecols='A:I', header=None)

    # Define the new header titles
    header_titles = ['품명', '모델/색상', '두께(㎜)', '폭(㎜)', '길이(㎜)', '전월재고(장)', '전월재고(㎡)', '현재고(장)', '현재고(㎡)']

    # Drop rows 0 to 3 (1행, 2행, 3행, 4행)
    df = df.drop(index=[0, 1, 2, 3])

    # Fill merged cells with the same data
    df = df.ffill(axis=0)

    # Set the header
    df.columns = header_titles

    # Drop rows containing '합     계' or '(이월작업 수식 복사용 행)'
    df = df[~df.apply(lambda row: row.astype(str).str.contains('합     계|\(이월작업 수식 복사용 행\)').any(), axis=1)]

    # C:F, H 열 데이터 숫자로 변환 및 천단위 콤마 추가 (G, I 열 제외)
    for col in df.columns[2:6].tolist() + [df.columns[7]]:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
        df[col] = df[col].apply(lambda x: f'{int(x):,}' if pd.notnull(x) else x)

    # G, I 열 데이터 소수점 0자리에서 반올림 후 천단위 콤마 추가
    for col in ['전월재고(㎡)', '현재고(㎡)']:
        df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', ''), errors='coerce')
        df[col] = df[col].apply(lambda x: f'{int(round(x)):,}' if pd.notnull(x) else x)

    # Save the dataframe to the final Excel file with the sheet name '잔여판재'
    with pd.ExcelWriter(final_file_path, mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name='잔여판재')

    # Apply additional formatting and printing options to the final file
    wb = openpyxl.load_workbook(final_file_path)
    ws = wb['잔여판재']

    # Set the header row
    for col_num, value in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=value)

    # 틀고정: 첫 번째 데이터 행 이후
    ws.freeze_panes = 'A2'

    # 열너비: 데이터 최대 너비에 맞게 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Adjust the column widths from C to the end to 12
    for col in ws.iter_cols(min_col=3):
        for cell in col:
            ws.column_dimensions[cell.column_letter].width = 12

    # 숫자 데이터: 천단위 구분자와 정수 표시 (G, I 열 제외)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=6):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.replace(',', '').isdigit():
                cell.value = int(cell.value.replace(',', ''))
                cell.number_format = '#,##0'

    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.replace(',', '').isdigit():
                cell.value = int(cell.value.replace(',', ''))
                cell.number_format = '#,##0'

    # G, I 열 데이터: 천단위 구분자와 정수 표시
    for col in [7, 9]:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.replace(',', '').replace('.', '').isdigit():
                    cell.value = int(float(cell.value.replace(',', '')))
                    cell.number_format = '#,##0'

    # 인쇄옵션 설정: 자동
    ws.print_title_rows = '1:1'  # 제목줄 반복
    ws.page_setup.fitToWidth = 1  # 가로폭 용지 맞춤
    ws.page_setup.orientation = 'portrait'  # 인쇄방향: 세로

    # 페이지 설정: 자동 설정
    ws.page_margins = openpyxl.worksheet.page.PageMargins(left=0.3937, right=0.3937, top=0.5906, bottom=0.5906, header=0.3937, footer=0.3937)

    # 머리글에 시트이름 넣기, 글자크기 20, 글씨체 맑은고딕
    ws.oddHeader.center.text = "&\"맑은 고딕,Bold\"&20 " + ws.title

    # 꼬리글 설정: 중간에 페이지/전체페이지, 오른쪽에 오늘 날짜/시간 표시
    ws.oddFooter.center.text = "&P/&N"
    ws.oddFooter.right.text = "&D &T"

    # Save the final formatted file
    wb.save(final_file_path)

    print(f'File saved to {final_file_path}')

if __name__ == "__main__":
    main()