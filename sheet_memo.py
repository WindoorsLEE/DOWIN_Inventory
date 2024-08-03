def main():
    # 기존 코드

    import pandas as pd
    from openpyxl import load_workbook

    # Load the Excel file
    file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-rmpw.xlsx'
    sheet_name = '원판재고'

    # Load workbook and select the sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Read the data from the specified sheet, starting from the 6th row to include data from 7th row
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=5)

    # Create a list to store memo comments
    memos = []

    # Loop through the cells in the J column (index 9, which corresponds to column 'J')
    for row in range(7, ws.max_row + 1):
        cell = ws[f'J{row}']
        if cell.comment:
            memo = cell.comment.text.replace('\n', ' ')
        else:
            memo = ''
        memos.append(memo)

    # Create a DataFrame for the memos
    memo_df = pd.DataFrame(memos, columns=['메모'])

    # Save the memo DataFrame to a new Excel file
    output_file_path = 'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-메모사항.xlsx'
    memo_df.to_excel(output_file_path, sheet_name='메모사항', index=False)

    print(f"파일이 저장되었습니다: {output_file_path}")


if __name__ == "__main__":
    main()
