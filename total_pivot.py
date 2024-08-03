def main():
    # 기존 코드

    import os
    import openpyxl
    from openpyxl import load_workbook

    # Directory path
    directory_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/'

    # Load the three Excel files
    file1_path = os.path.join(directory_path, '재고수불부-sheet_pivot2data.xlsx')
    file2_path = os.path.join(directory_path, '재고수불부-submaterial_pivot.xlsx')
    file3_path = os.path.join(directory_path, '재고수불부-partial_pivot.xlsx')

    # Load the workbooks
    wb1 = load_workbook(file1_path)
    wb2 = load_workbook(file2_path)
    wb3 = load_workbook(file3_path)

    # Load the sheets
    ws1 = wb1['원판재고']
    ws2 = wb2['부자재']
    ws3 = wb3['잔여판재']

    # Create a new workbook
    new_wb = openpyxl.Workbook()

    # Remove the default sheet created
    new_wb.remove(new_wb.active)

    # Function to copy the sheet with all settings
    def copy_sheet(source_ws, target_wb, sheet_name, set_col_widths=False):
        target_ws = target_wb.create_sheet(title=sheet_name)
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell.font = openpyxl.styles.Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        vertAlign=cell.font.vertAlign,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color
                    )
                    new_cell.border = openpyxl.styles.Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom,
                        diagonal=cell.border.diagonal,
                        diagonal_direction=cell.border.diagonal_direction,
                        outline=cell.border.outline,
                        vertical=cell.border.vertical,
                        horizontal=cell.border.horizontal
                    )
                    new_cell.fill = openpyxl.styles.PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color
                    )
                    new_cell.number_format = cell.number_format
                    new_cell.protection = openpyxl.styles.Protection(
                        locked=cell.protection.locked,
                        hidden=cell.protection.hidden
                    )
                    new_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical='center',
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent
                    )
        if set_col_widths:
            column_widths = {
                'A': 8, 'B': 12, 'C': 5, 'D': 7,
                'E': 2, 'F': 2, 'G': 2, 'H': 2, 'I': 2, 'J': 2, 'K': 2, 'L': 2, 'M': 2, 'N': 2, 'O': 2, 'P': 2, 'Q': 2, 'R': 2,
                'S': 9, 'T': 10, 'U': 10, 'V': 10,
                'W': 2, 'X': 2, 'Y': 2,
                'Z': 9, 'AA': 9, 'AB': 9, 'AC': 9,
                'AD': 8, 'AG': 10, 'AH': 10
            }
            for col, width in column_widths.items():
                target_ws.column_dimensions[col].width = width

            # 자동 줄바꿈 설정 및 세로방향 가운데 맞춤 설정
            for row in target_ws.iter_rows():
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')

        for col_dim in source_ws.column_dimensions:
            target_ws.column_dimensions[col_dim].width = source_ws.column_dimensions[col_dim].width

        for row_dim in source_ws.row_dimensions:
            target_ws.row_dimensions[row_dim].height = source_ws.row_dimensions[row_dim].height

        target_ws.print_title_rows = source_ws.print_title_rows
        target_ws.page_setup.orientation = source_ws.page_setup.orientation
        target_ws.page_setup.fitToWidth = source_ws.page_setup.fitToWidth
        target_ws.page_setup.fitToHeight = source_ws.page_setup.fitToHeight
        target_ws.page_margins = source_ws.page_margins
        target_ws.oddHeader.center.text = source_ws.oddHeader.center.text
        target_ws.oddFooter.center.text = source_ws.oddFooter.center.text
        target_ws.oddFooter.right.text = source_ws.oddFooter.right.text

    # Copy sheets with all settings
    copy_sheet(ws1, new_wb, '원판재고')
    copy_sheet(ws2, new_wb, '부자재')
    copy_sheet(ws3, new_wb, '잔여판재')

    # Save the new workbook
    output_file_path = os.path.join(directory_path, '재고수불부-total_pivot.xlsx')
    new_wb.save(output_file_path)

    print(f"File saved to {output_file_path}")


if __name__ == "__main__":
    main()
