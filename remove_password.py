def main():
    # 기존 코드

    import shutil
    import os
    import msoffcrypto
    import openpyxl

    # 원본 파일 경로
    original_file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부.xlsx'
    # 비밀번호가 제거된 중간 파일 경로
    intermediate_file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-rmpw.xlsx'
    # 현재 비밀번호
    password = 'endls6688'

    # 비밀번호가 제거된 파일 생성
    with open(original_file_path, 'rb') as file:
        decrypted = msoffcrypto.OfficeFile(file)
        decrypted.load_key(password=password)
        with open(intermediate_file_path, 'wb') as decrypted_file:
            decrypted.decrypt(decrypted_file)

    # 비밀번호가 제거된 파일을 엽니다.
    wb = openpyxl.load_workbook(intermediate_file_path, data_only=True)

    # 모든 시트의 수식을 값으로 변환
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 수식이 있는 셀
                    cell.value = cell.value

    # 변환된 파일을 다시 저장합니다.
    wb.save(intermediate_file_path)
    print(f"파일이 성공적으로 비밀번호가 제거되고 수식이 값으로 변환되어 '{intermediate_file_path}'에 저장되었습니다.")

if __name__ == "__main__":
    main()
