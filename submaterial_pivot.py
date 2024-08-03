def main():
    # 기존 코드

    import pandas as pd
    import openpyxl

    # Load the intermediate Excel file to get the second sheet
    intermediate_file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-rmpw.xlsx'
    final_file_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/재고수불부-submaterial_pivot.xlsx'

    # Define the replacements
    replacements = {
        '두께를 불문하고 가스켓마감바끼리는 되도록 같이 발주하지 말것(제품 분류시 혼란을 막기 위함임!)': '발주시 문의',
        '2016년 11월 15일에 입고. 이스라엘 제품은 맞으나 입고 정보는 알 수 없음 (탑갈 제품 적용은 가능한 것으로 보임)': '입고정보없음',
        'PC커넥터용(폴리유용 신형) - 현재 커넥터에 안맞음': '현재커넥터안맞음',
        '2023.8.28(월) 스크랩 처리 완료 : 대표이사 지시로 처리함': '스크랩 지시',
        '영주 가흥 주차타워(유길환팀장 주문사이즈)': '영주가흥',
        '상원이엔지 주문사이즈(송수근 본부장)': '상원이엔지',
        '무피막 및 기타도장색상 잔여분': '무피막 잔여분',
        'P C 커 넥 터': 'PC커넥터'
    }

    # Load the second sheet data
    df = pd.read_excel(intermediate_file_path, sheet_name=1, usecols='A:J', header=None)

    # Define the new header titles
    header_titles = ['품목코드', '범주', '품명', '구조', '색상/품명/규격', '길이(m,mm)/(EA)', '단위중량(kg/m)', '총무게(Kg)', '전월재고(EA)', '현재고(EA)']

    # Drop rows 0, 1, 5 (1행, 2행, 6행)
    df = df.drop(index=[0, 1, 5])

    # Fill merged cells with the same data except for columns G and H
    df.iloc[:, :6] = df.iloc[:, :6].ffill(axis=0)

    # Set the header
    df.columns = header_titles

    # Drop rows 3 to 5 (adjusted to account for previous drops)
    df = df.drop(index=[2, 3, 4])

    # Drop rows containing '소 계', 'AL총계', '합          계'
    df = df[~df.apply(lambda row: row.astype(str).str.contains('소 계|AL총계|합          계').any(), axis=1)]

    # Replace text based on replacements dictionary
    def replace_text(cell):
        if isinstance(cell, str):
            for key, value in replacements.items():
                cell = cell.replace(key, value)
        return cell

    df = df.applymap(replace_text)

    # Save the dataframe to the final Excel file with the sheet name '부자재'
    with pd.ExcelWriter(final_file_path, mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name='부자재')

    # Apply additional formatting and printing options to the final file
    wb = openpyxl.load_workbook(final_file_path)
    ws = wb['부자재']

    # Set the header row
    for col_num, value in enumerate(header_titles, 1):
        ws.cell(row=1, column=col_num, value=value)

    # F열: 소수점 1자리까지 표시
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.0'

    # G열: 소수점 4자리까지 표시
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.0000'

    # H열: 소수점 2자리까지 표시
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.00'

    # 틀고정: D2
    ws.freeze_panes = 'D2'

    # 열너비: 데이터 최대 너비에 맞게 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Adjust the column widths from F to the end to 16
    for col in ws.iter_cols(min_col=6, max_col=ws.max_column):
        for cell in col:
            ws.column_dimensions[cell.column_letter].width = 16

    # 인쇄옵션 설정: 자동
    ws.print_title_rows = '1:1'  # 제목줄 반복
    ws.page_setup.fitToWidth = 1  # 가로폭 용지 맞춤
    ws.page_setup.orientation = 'landscape'  # 인쇄방향: 가로

    # 페이지 설정: 자동 설정
    ws.page_margins = openpyxl.worksheet.page.PageMargins(left=0.3937, right=0.3937, top=0.5906, bottom=0.5906, header=0.3937, footer=0.3937)

    # 머리글에 시트이름 넣기, 글자크기 20, 글씨체 맑은고딕
    ws.oddHeader.center.text = "&\"맑은 고딕,Bold\"&20 " + ws.title

    # 꼬리글 설정: 중간에 페이지/전체페이지, 오른쪽에 오늘 날짜/시간 표시
    ws.oddFooter.center.text = "&P/&N"
    ws.oddFooter.right.text = "&D &T"

    # Save the final formatted file
    wb.save(final_file_path)

    print(f'File saved to {final_file_path}')


if __name__ == "__main__":
    main()
