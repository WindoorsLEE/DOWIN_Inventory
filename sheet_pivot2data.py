def main():
    # 기존 코드

    import pandas as pd
    import os
    import openpyxl

    # 파일 위치 폴더 및 파일 경로
    folder_path = r'C:/Users/windo/OneDrive/바탕 화면/Zenbook2/DOWIN/inventory/'
    source_file_path = os.path.join(folder_path, '재고수불부-sheet_pivot.xlsx')
    destination_file_path = os.path.join(folder_path, '재고수불부-sheet_pivot2data.xlsx')

    # 파일 읽기
    df = pd.read_excel(source_file_path, sheet_name='원판재고')

    # 열 이름 출력하여 확인
    print("Columns before processing:")
    print(df.columns)

    # 열 병합 함수
    def concatenate_columns(df, new_col, col_list):
        # 1 순위: 모든 데이터를 소수점 0자리로 반올림
        for col in col_list:
            df[col] = df[col].apply(lambda x: round(x, 0) if isinstance(x, (int, float)) else x)
        
        # 2 순위: NaN 값을 ""로 대체
        df[col_list] = df[col_list].fillna("")

        # 3 순위: 셀 병합 시 셀 간 데이터 중간에 " / " 삽입하여 병합
        df[new_col] = df[col_list].astype(str).agg(' / '.join, axis=1)

        # 4 순위: "입고예정" 열의 병합 데이터에서 소수점 표기를 제거
        df[new_col] = df[new_col].str.replace(r'\.0', '', regex=True)

        # 5 순위: 빈 셀 병합 시 " / " 제거
        df[new_col] = df[new_col].apply(lambda x: ' / '.join([i for i in x.split(' / ') if i.strip()]))

        df.drop(columns=col_list, inplace=True)

    # 정확한 열 이름을 사용하여 병합 작업 수행
    concatenate_columns(df, '입고예정', ['입고예정일', '입고(장)', '입고(㎡)', '입항예정일'])
    concatenate_columns(df, '서울당월출고예정', ['서울당월현장', '서울당월수량(㎡)'])
    concatenate_columns(df, '서울+1월출고예정', ['서울+1월현장', '서울+1월수량(㎡)'])
    concatenate_columns(df, '서울+2월출고예정', ['서울+2월현장', '서울+2월수량(㎡)'])
    concatenate_columns(df, '대구당월출고예정', ['대구당월현장', '대구당월수량(㎡)'])
    concatenate_columns(df, '대구+1월출고예정', ['대구+1월현장', '대구+1월수량(㎡)'])
    concatenate_columns(df, '대구+2월출고예정', ['대구+2월현장', '대구+2월수량(㎡)'])
    concatenate_columns(df, 'REMARK_메모', ['REMARK', '메모'])

    # 'REMARK_메모'를 마지막 열로 이동
    df['REMARK_메모'] = df.pop('REMARK_메모')

    # 정확한 열 이름 확인
    print("Columns after concatenation and moving:")
    print(df.columns)

    # NaN 값 삭제
    df.dropna(inplace=True)

    # 열 이름 출력하여 확인
    print("Columns after processing:")
    print(df.columns)

    # 숫자를 1000단위로 구분하고 소수점 반올림하여 정수로 표기
    numeric_columns = df.select_dtypes(include=['float64', 'int64']).columns
    df[numeric_columns] = df[numeric_columns].applymap(lambda x: int(round(x)) if pd.notna(x) else x)

    # 파일 저장
    with pd.ExcelWriter(destination_file_path) as writer:
        df.to_excel(writer, index=False, sheet_name='원판재고')

    # 저장된 파일을 열어 열 너비 자동 조정 및 틀 고정 설정
    wb = openpyxl.load_workbook(destination_file_path)
    ws = wb['원판재고']

    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # D:K 열 숫자 형식 설정 및 오른쪽 정렬
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

    # J 열 음수 값 빨간색 설정
    for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = openpyxl.styles.Font(color="FF0000")
                cell.number_format = '#,##0'
                cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')

    # A:J 열 자동 줄바꿈 설정
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')

    # K 열부터 끝까지 열 너비 16으로 설정
    for col in ws.iter_cols(min_col=11):
        for cell in col:
            ws.column_dimensions[cell.column_letter].width = 16

    # 모든 셀 자동 줄바꿈 설정 및 세로 방향 가운데 정렬
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')

    # 제목줄 셀 가운데 맞춤
    for cell in ws[1]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 열너비 조정
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 32
    for col in range(4, 11):  # D to J
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    for col in range(11, ws.max_column + 1):  # K to end
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # 틀 고정 설정
    ws.freeze_panes = ws['D2']

    # 파일 저장
    wb.save(destination_file_path)

    print(f"File saved to {destination_file_path}")


if __name__ == "__main__":
    main()
